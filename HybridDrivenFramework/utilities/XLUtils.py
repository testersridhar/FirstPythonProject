import openpyxl


# Read rowcount
def gerRowCount(fileName, sheetName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    return sheet.max_row

# Read column count
def gerColumnCount(fileName, sheetName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    return sheet.max_column


# Read data from cell from the sheet
def readData(fileName, sheetName, rowCount, columnCount):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowCount, column=columnCount).value


# Write data to excelsheet
def write(fileName, sheetName, rowCount, columnCount, data):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    sheet.cell(row=rowCount, column=columnCount).value = data
    workbook.save(fileName)
